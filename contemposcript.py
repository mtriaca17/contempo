import openpyxl
import numpy as np
from openpyxl.utils import get_column_letter, column_index_from_string

wb = openpyxl.load_workbook('AAPL.xlsx', data_only=True)
sheet = wb.get_sheet_by_name('AAPL')
dataArray = []
dataClean = []
textFile = open('text.txt', 'w')

for row in range(2, sheet.max_row+1):
	for col in range(1, 7):
		data = 'insertData({0},"{1}", {2}, {3},{4},{5},{6},{7});'.format(row-1, sheet.cell(row=row, column = col).value, sheet.cell(row=row, column = col+1).value,sheet.cell(row=row, column = col+2).value,sheet.cell(row=row, column = col+3).value,sheet.cell(row=row, column = col+4).value,sheet.cell(row=row, column = col+5).value,sheet.cell(row=row, column = col+6).value)
		#print(data)
		dataArray.append(data)
		#print(dataArray
		break
#listClean(dataArray)

for item in dataArray:
	textFile.write("%s\n" % item)